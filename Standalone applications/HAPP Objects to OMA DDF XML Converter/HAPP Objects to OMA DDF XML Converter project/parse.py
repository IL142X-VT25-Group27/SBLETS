##############################################

# 2024-06-18 Alexander Ström

# This file generates an OMA DDF xml file from HAPP LwM2M Object Specification xlxs (Excel). Developed to work with
# structure used in HAPP002 Version: P and should work with Resources sheet that uses the same structure. This is
# then imported into Leshan demo modelsfolder -m.

xlsxFileName = "LwM2M_objects.xlsx"  # Change input xlsx file name here if needed
firstID = 27002  # First non registered object ID in xlsx from top to bottom (important for correct parsing)

##############################################

import sys, io

buffer = io.StringIO()
sys.stdout = sys.stderr = buffer

from openpyxl import load_workbook
import pandas as pd
import xml.etree.ElementTree as minidom
from lxml import etree
import eel
import base64
import os

firstID = firstID - 1 # Fix

# Initialize the Eel app
eel.init('ui')

# Upload file from UI
@eel.expose
def uploadFile(filename, filedata):
    # Remove the base64 header
    header, encoded = filedata.split(",", 1)
    # Decode the base64 string
    data = base64.b64decode(encoded)

    # Define the path to save the file
    savePath = os.path.join("uploads", filename)
    os.makedirs(os.path.dirname(savePath), exist_ok=True)

    # Write the file to the filesystem
    with open(savePath, "wb") as f:
        f.write(data)

    parseAndCreate(savePath)

    savedToPath = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ddf')

    return f"Sucessfully generted and saved XML files from '{filename}' to path '{savedToPath}'"

def parseAndCreate(filePath):
    xlsxFileName = filePath

    wb = load_workbook(filename=xlsxFileName, read_only=True)
    ws = wb['Resources']
    wsObjects = wb['Objects']
    # Get data from Resources sheet
    lastRow = 'L' + str(ws.max_row)
    dataRows = []
    for row in ws['B4':lastRow]:
        dataCols = []
        for cell in row:
            dataCols.append(cell.value)
        dataRows.append(dataCols)
    # Array to dataframe
    df = pd.DataFrame(dataRows)
    # Set the first row as the header and remove it from the data
    df.columns = df.iloc[0]
    df = df[1:]
    df.reset_index(drop=True, inplace=True)
    # Get data from Objects sheet
    lastRowO = 'L' + str(wsObjects.max_row)
    dataRowsO = []
    for row in wsObjects['B4':lastRowO]:
        dataColsO = []
        for cell in row:
            dataColsO.append(cell.value)
        dataRowsO.append(dataColsO)
    # Array to dataframe
    dfObjects = pd.DataFrame(dataRowsO)
    # Set the first row as the header and remove it from the data
    dfObjects.columns = dfObjects.iloc[0]
    dfObjects = dfObjects[1:]
    dfObjects.reset_index(drop=True, inplace=True)
    filteredDf = df[df['Object ID'] > firstID]  # First non registered object ID
    uniqueIDs = filteredDf['Object ID'].unique()
    groupedObjects = {}
    # Groups objects with same ID to make it easier to parse
    for objectID in uniqueIDs:
        groupDf = df[df['Object ID'] == objectID]
        groupedObjects[objectID] = groupDf.to_dict(orient='records')
    for objectID, group in groupedObjects.items():
        # Roots for XML file
        root = minidom.Element('LWM2M')
        root.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
        root.set("xsi:noNamespaceSchemaLocation", "LWM2M.xsd")
        mainObject = minidom.Element('Object')
        mainObject.set("ObjectType", "MODefinition")
        root.append(mainObject)

        firstRow = group[0]

        # Get version from Objects sheet
        versionRow = dfObjects[dfObjects['Name'] == firstRow["Object Name"]]
        version = versionRow['Object Version'].iloc[0]

        # Object universal elements
        print("Building object for", firstRow["Object Name"], "version", version)
        logData = "Building object for", firstRow["Object Name"], "version", version
        eel.addToLog(str(logData))
        Name = minidom.SubElement(mainObject, "Name")
        Name.text = firstRow["Object Name"]
        Description1 = minidom.SubElement(mainObject, "Description1")
        Description1.text = " "
        ObjectID = minidom.SubElement(mainObject, "ObjectID")
        ObjectID.text = str(firstRow["Object ID"])
        ObjectURN = minidom.SubElement(mainObject, "ObjectURN")
        ObjectURN.text = (f'urn:oma:lwm2m:x:{firstRow["Object ID"]}:{version}')
        ObjectVersion = minidom.SubElement(mainObject, "ObjectVersion")
        ObjectVersion.text = version
        MultipleInstances = minidom.SubElement(mainObject, "MultipleInstances")
        MultipleInstances.text = "Multiple"
        Mandatory = minidom.SubElement(mainObject, "Mandatory")
        Mandatory.text = "Mandatory"

        Resources = minidom.Element('Resources')
        mainObject.append(Resources)

        # Get object parameters and make xml element item
        for row in group:
            print(row)
            Item = minidom.Element("Item")
            Item.set("ID", f'{row["Resource ID"]}')
            Resources.append(Item)

            Name = minidom.SubElement(Item, "Name")
            Name.text = row["Resource Name"]
            Operations = minidom.SubElement(Item, "Operations")
            Operations.text = row["Operations"]
            MultipleInstances = minidom.SubElement(Item, "MultipleInstances")
            MultipleInstances.text = row["Instances"]
            Mandatory = minidom.SubElement(Item, "Mandatory")
            Mandatory.text = row["Mandatory"]
            Type = minidom.SubElement(Item, "Type")
            Type.text = row["Type"]
            RangeEnumeration = minidom.SubElement(Item, "RangeEnumeration")
            RangeEnumeration.text = row["Range or Enumeration"]
            Units = minidom.SubElement(Item, "Units")
            Units.text = row["Units"]
            Description = minidom.SubElement(Item, "Description")
            Description.text = row["Description"]

        Description2 = minidom.SubElement(mainObject, "Description2")
        Description2.text = " "

        # Create tree from content
        tree = minidom.ElementTree(root)

        # Name with no blank spaces
        nameNonSpace = firstRow["Object Name"].replace(" ", "_")
        fileName = f"{firstRow['Object ID']}_{nameNonSpace}_v{version}.xml"

        ddfPath = os.path.join('ddf', fileName)
        os.makedirs(os.path.dirname(ddfPath), exist_ok=True)

        # Saving raw structure to file
        print("Saving file for", fileName)
        with open(ddfPath, "wb") as file:
            tree.write(file)

        # Pretty print XML file
        rawFile = etree.parse(ddfPath)
        pretty_xml = etree.tostring(rawFile, pretty_print=True, encoding=str)
        with open(ddfPath, "w", encoding="utf-8") as file:
            file.write(pretty_xml)

eel.start('index.html', size=(800, 600))