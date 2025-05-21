##################### -- generateOmaDdf.py -- #########################

# 2024-06-18 Alexander Ström
# This component generates multiple OMA DDF xml file from HAPP LwM2M Object Specification xlxs (Excel). Developed to work with
# structure used in HAPP002 Version: P and should work with a Resources sheet that uses the same structure. This is
# then imported into Leshan demo modelsfolder -m.

xlsxFileName = "LwM2M_objects.xlsx"  # Change input xlsx file name here if needed
firstID = 27002  # First non registered object ID in xlsx from top to bottom (important for correct parsing) - First object to parse
# username = "user" # Username for a directory path

##############################################################

# chromium-browser must be installed on machine

from openpyxl import load_workbook
import pandas as pd
import xml.etree.ElementTree as ET
from xml.dom import minidom
from lxml import etree
import eel
import base64
import os
import configparser
import logging

firstID = firstID - 1  # Fix

# Create a ConfigParser object
config = configparser.ConfigParser()

# Read the configuration file
config.read('config.ini')

leshanPath = config.get('LESHAN', 'Leshan objects path')

logging.debug(f"Leshan objects path: {leshanPath}")

# Create a ConfigParser object
config = configparser.ConfigParser()

# Read the configuration file
config.read('config.ini')

# Upload file from UI
@eel.expose
def uploadFile(filename, filedata):
    # Remove the base64 header
    header, encoded = filedata.split(",", 1)
    # Decode the base64 string
    data = base64.b64decode(encoded)

    # Define the path to save the file
    savePath = os.path.join("uploads", filename)
    os.makedirs(os.path.dirname(savePath), exist_ok=True)

    # Write the file to the filesystem
    with open(savePath, "wb") as f:
        f.write(data)

    parseAndCreate(savePath)

    # savedToPath = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ddf')

    logging.info(f"Sucessfully generted and saved XML files from '{filename}' to path '{leshanPath}'")
    return f"Sucessfully generted and saved XML files from '{filename}' to path '{leshanPath}'"


def parseAndCreate(filePath):
    xlsxFileName = filePath

    wb = load_workbook(filename=xlsxFileName, read_only=True)
    ws = wb['Resources']
    wsObjects = wb['Objects']
    # Get data from a resources sheet
    lastRow = 'L' + str(ws.max_row)
    dataRows = []
    for row in ws['B4':lastRow]:
        dataCols = []
        for cell in row:
            dataCols.append(cell.value)
        dataRows.append(dataCols)
    # Array to dataframe
    df = pd.DataFrame(dataRows)
    # Set the first row as the header and remove it from the data
    df.columns = df.iloc[0]
    df = df[1:]
    df.reset_index(drop=True, inplace=True)
    # Get data from an objects sheet
    lastRowO = 'L' + str(wsObjects.max_row)
    dataRowsO = []
    for row in wsObjects['B4':lastRowO]:
        dataColsO = []
        for cell in row:
            dataColsO.append(cell.value)
        dataRowsO.append(dataColsO)
    # Array to dataframe
    dfObjects = pd.DataFrame(dataRowsO)
    # Set the first row as the header and remove it from the data
    dfObjects.columns = dfObjects.iloc[0]
    dfObjects = dfObjects[1:]
    dfObjects.reset_index(drop=True, inplace=True)
    filteredDf = df[df['Object ID'] > firstID]  # First non registered object ID
    uniqueIDs = filteredDf['Object ID'].unique()
    groupedObjects = {}
    # Groups objects with the same ID to make it easier to parse
    for objectID in uniqueIDs:
        groupDf = df[df['Object ID'] == objectID]
        groupedObjects[objectID] = groupDf.to_dict(orient='records')
    for objectID, group in groupedObjects.items():
        # Roots for XML file
        root = ET.Element('LWM2M')
        root.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
        root.set("xsi:noNamespaceSchemaLocation", "LWM2M.xsd")
        mainObject = ET.Element('Object')
        mainObject.set("ObjectType", "MODefinition")
        root.append(mainObject)

        firstRow = group[0]

        # Get version from Objects sheet
        versionRow = dfObjects[dfObjects['Name'] == firstRow["Object Name"]]
        version = versionRow['Object Version'].iloc[0]

        # Object universal elements
        logging.info(f"Building object for {firstRow['Object Name']} version {version}")
        print("Building object for", firstRow["Object Name"], "version", version)
        logData = "Building object for", firstRow["Object Name"], "version", version
        eel.addToLog(str(logData), "modelGen")
        Name = ET.SubElement(mainObject, "Name")
        Name.text = firstRow["Object Name"]
        Description1 = ET.SubElement(mainObject, "Description1")
        Description1.text = " "
        ObjectID = ET.SubElement(mainObject, "ObjectID")
        ObjectID.text = str(firstRow["Object ID"])
        ObjectURN = ET.SubElement(mainObject, "ObjectURN")
        ObjectURN.text = f'urn:oma:lwm2m:x:{firstRow["Object ID"]}:{version}'
        ObjectVersion = ET.SubElement(mainObject, "ObjectVersion")
        ObjectVersion.text = version
        MultipleInstances = ET.SubElement(mainObject, "MultipleInstances")
        MultipleInstances.text = "Multiple"
        Mandatory = ET.SubElement(mainObject, "Mandatory")
        Mandatory.text = "Mandatory"

        Resources = ET.Element('Resources')
        mainObject.append(Resources)

        # Get object parameters and make xml element item
        for row in group:
            print(row)
            Item = ET.Element("Item")
            Item.set("ID", f'{row["Resource ID"]}')
            Resources.append(Item)

            Name = ET.SubElement(Item, "Name")
            Name.text = row["Resource Name"]
            Operations = ET.SubElement(Item, "Operations")
            Operations.text = row["Operations"]
            MultipleInstances = ET.SubElement(Item, "MultipleInstances")
            MultipleInstances.text = row["Instances"]
            Mandatory = ET.SubElement(Item, "Mandatory")
            Mandatory.text = row["Mandatory"]
            Type = ET.SubElement(Item, "Type")
            Type.text = row["Type"]
            RangeEnumeration = ET.SubElement(Item, "RangeEnumeration")
            RangeEnumeration.text = row["Range or Enumeration"]
            Units = ET.SubElement(Item, "Units")
            Units.text = row["Units"]
            Description = ET.SubElement(Item, "Description")
            Description.text = row["Description"]

        Description2 = ET.SubElement(mainObject, "Description2")
        Description2.text = " "

        # Create a tree from content
        tree = ET.ElementTree(root)

        # Name with no blank spaces
        nameNonSpace = firstRow["Object Name"].replace(" ", "_")
        fileName = f"{firstRow['Object ID']}_{nameNonSpace}_v{version}.xml"

        xmlString = ET.tostring(root, encoding='unicode')

        dom = minidom.parseString(xmlString)
        prettyXml = dom.toprettyxml(indent="  ")

        eel.download(fileName, prettyXml) # Send XML file for download on eel app

        ddfPath = os.path.join(leshanPath, fileName)  # dir for SBLETS server
        os.makedirs(os.path.dirname(ddfPath), exist_ok=True)

        # Saving raw structure to file
        if leshanPath != "False":
            try:
                print("Saving file for", fileName)
                logging.info(f"Leshan objects path: {fileName}")
                with open(ddfPath, "wb") as file:
                    tree.write(file)

                # Pretty print XML file
                rawFile = etree.parse(ddfPath)
                pretty_xml = etree.tostring(rawFile, pretty_print=True, encoding=str)
                with open(ddfPath, "w", encoding="utf-8") as file:
                    file.write(pretty_xml)
            except Exception as e:
                logging.error(e)
                pass